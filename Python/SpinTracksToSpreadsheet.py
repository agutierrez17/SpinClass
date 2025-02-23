import pandas as pd
import pyodbc
import warnings
from openpyxl import load_workbook
import numpy as np

warnings.filterwarnings("ignore")

# Connect to database and open SQL cursor
print('Connecting to database...')
print('')
conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};'
                      'Trusted_Connection=yes;')
cursor = conn.cursor()

# Query Tracks to Excel view
print('Querying data from Tracks to Excel view...')
print('')
sql = """
SELECT
[External URL],
[Category],
[Builder],
[Ride Code],
[Position],
[Track Name],
[Artists],
[BPM],
[Used RPM],
[Duration],
[Genre],
[Notes]
FROM [SpinClass].[dbo].[ExcelTracksView]
"""
df = pd.read_sql(sql,conn)

# Open up pandas ExcelWriter
print('Opening up SongLibrary.xlsx Excel sheet...')
print('')
book = load_workbook(path)
sheet_name='Song Library & Playlist Builder'
sheet = book[sheet_name]
last_row = sheet.max_row

# Write data to SongLibrary Excel sheet
print('Writing track data to song library...')
print('')
try: 
    for row in df.values.tolist():
        sheet.append(row)
    print('All tracks written to spreadsheet.')
    print('')
except:
    print('Error: new tracks not written to spreadsheet')
    print('')

book.save(path)

# Read SongLibrary excel data into dataframe
print('Refreshing database table...')
print('')
data = pd.read_excel(path, skiprows=6, index_col=None,converters={'Builder':str,'Ride Code':str,'Position':str,'Name':str,'Notes':str,'BPM':float,'Used RPM':float})
data = data.fillna("")

# Truncate SongLibrary table
cursor.execute("""TRUNCATE TABLE [dbo].[SongLibrary]""")
cursor.commit()

# Insert data into SongLibrary table
cursor.executemany("""
INSERT INTO [dbo].[SongLibrary]([External URL],[Category],[Builder],[Ride Code],[Position],[Track Name],[Artists],[BPM],[Used RPM],[Duration],[Genre],[Notes])
     VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""", data.values.tolist())
cursor.commit()

print('All tracks inserted.')
print('')

cursor.close()
